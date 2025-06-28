# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import xlrd
import xlsxwriter
import base64
import io
from datetime import datetime
import logging

_logger = logging.getLogger(__name__)


class BankTransactionImportWizard(models.TransientModel):
    _name = 'dpt.bank.transaction.import.wizard'
    _description = 'Wizard Import Giao Dịch Ngân Hàng'

    name = fields.Char('Tên Import', default='Import Giao Dịch Ngân Hàng')
    excel_file = fields.Binary('File Excel',
                               help="Upload file Excel chứa dữ liệu giao dịch ngân hàng")
    filename = fields.Char('Tên File')

    # Cấu hình mapping cột
    date_column = fields.Integer('Cột Ngày Giao Dịch', default=1,
                                 help="Số thứ tự cột chứa ngày giao dịch (bắt đầu từ 1)")
    description_column = fields.Integer('Cột Mô Tả', default=2,
                                        help="Số thứ tự cột chứa mô tả giao dịch")
    amount_column = fields.Integer('Cột Số Tiền', default=3,
                                   help="Số thứ tự cột chứa số tiền")
    reference_column = fields.Integer('Cột Mã Tham Chiếu', default=4,
                                      help="Số thứ tự cột chứa mã tham chiếu")
    balance_column = fields.Integer('Cột Số Dư', default=5,
                                    help="Số thứ tự cột chứa số dư sau giao dịch")

    # Cấu hình import
    start_row = fields.Integer('Dòng Bắt Đầu', default=2,
                               help="Dòng bắt đầu đọc dữ liệu (bỏ qua header)")
    date_format = fields.Selection([
        ('dd/mm/yyyy', 'DD/MM/YYYY'),
        ('mm/dd/yyyy', 'MM/DD/YYYY'),
        ('yyyy-mm-dd', 'YYYY-MM-DD'),
        ('dd-mm-yyyy', 'DD-MM-YYYY'),
    ], string='Định Dạng Ngày', default='dd/mm/yyyy')

    account_id = fields.Many2one('dpt.fund.account', 'Tài Khoản Ngân Hàng', required=True,
                                 help="Tài khoản ngân hàng sẽ được cập nhật giao dịch")

    # Kết quả import
    import_status = fields.Selection([
        ('draft', 'Chưa Import'),
        ('preview', 'Xem Trước'),
        ('done', 'Hoàn Thành'),
        ('error', 'Có Lỗi')
    ], default='draft', string='Trạng Thái')

    total_rows = fields.Integer('Tổng Số Dòng', readonly=True)
    success_count = fields.Integer('Số Dòng Thành Công', readonly=True)
    error_count = fields.Integer('Số Dòng Lỗi', readonly=True)
    error_log = fields.Text('Log Lỗi', readonly=True)

    # Preview data
    preview_data = fields.Text('Dữ Liệu Xem Trước', readonly=True)

    @api.model
    def default_get(self, fields_list):
        """Set default account if coming from account view"""
        res = super().default_get(fields_list)
        if self.env.context.get('active_model') == 'dpt.fund.account':
            res['account_id'] = self.env.context.get('active_id')
        return res

    @api.model
    def _check_excel_dependencies(self):
        """Kiểm tra dependencies Excel"""
        if not xlrd:
            raise UserError(_('Thiếu thư viện xlrd. Vui lòng cài đặt: pip install xlrd'))
        if not xlsxwriter:
            raise UserError(_('Thiếu thư viện xlsxwriter. Vui lòng cài đặt: pip install xlsxwriter'))

    def action_preview_data(self):
        """Xem trước dữ liệu từ file Excel"""
        self._check_excel_dependencies()

        if not self.excel_file:
            raise UserError(_('Vui lòng upload file Excel'))

        try:
            # Đọc file Excel với encoding tốt hơn
            file_data = base64.b64decode(self.excel_file)

            # Hỗ trợ cả .xls và .xlsx
            if self.filename and self.filename.endswith('.xlsx'):
                # Sử dụng openpyxl cho .xlsx
                try:
                    from openpyxl import load_workbook
                    workbook = load_workbook(io.BytesIO(file_data))
                    worksheet = workbook.active
                    return self._process_xlsx_preview(worksheet)
                except ImportError:
                    raise UserError(_('Thiếu thư viện openpyxl cho file .xlsx. Vui lòng cài đặt: pip install openpyxl'))
            else:
                # Sử dụng xlrd cho .xls
                workbook = xlrd.open_workbook(file_contents=file_data)
                worksheet = workbook.sheet_by_index(0)
                return self._process_xlsx_preview(worksheet)

        except Exception as e:
            raise UserError(_('Lỗi đọc file Excel: %s') % str(e))

    def _process_xlsx_preview(self, worksheet):
        """Xử lý preview cho file .xlsx"""
        preview_lines = []
        error_lines = []

        # Đọc tối đa 10 dòng để preview
        max_preview = min(self.start_row + 9, worksheet.max_row + 1)

        for row_idx in range(self.start_row, max_preview):
            try:
                row_data = self._parse_xlsx_row(worksheet, row_idx)
                if row_data:
                    preview_lines.append(
                        f"Dòng {row_idx}: {row_data['date']} | "
                        f"{row_data['description']} | "
                        f"{row_data['amount']:,.0f} VND | "
                        f"Ref: {row_data['reference']}"
                    )
            except Exception as e:
                error_lines.append(f"Dòng {row_idx}: Lỗi - {str(e)}")

        return self._create_preview_response(preview_lines, error_lines, worksheet.max_row)

    def _create_preview_response(self, preview_lines, error_lines, total_rows):
        """Tạo response cho preview"""
        preview_text = "=== DỮ LIỆU XEM TRƯỚC ===\n"
        preview_text += f"Tài khoản: {self.account_id.name}\n"
        preview_text += f"Định dạng ngày: {self.date_format}\n"
        preview_text += f"Tổng số dòng trong file: {total_rows}\n"
        preview_text += f"Dòng bắt đầu: {self.start_row}\n\n"

        if preview_lines:
            preview_text += "--- DÒNG HỢP LỆ ---\n"
            preview_text += "\n".join(preview_lines)

        if error_lines:
            preview_text += "\n\n--- DÒNG CÓ LỖI ---\n"
            preview_text += "\n".join(error_lines)

        self.write({
            'preview_data': preview_text,
            'import_status': 'preview',
            'total_rows': total_rows - (self.start_row - 1)
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context
        }

    def action_import_transactions(self):
        """Thực hiện import giao dịch"""
        if not self.excel_file:
            raise UserError(_('Vui lòng upload file Excel'))

        try:
            # Đọc file Excel
            file_data = base64.b64decode(self.excel_file)
            workbook = xlrd.open_workbook(file_contents=file_data)
            worksheet = workbook.sheet_by_index(0)

            success_count = 0
            error_count = 0
            error_messages = []
            created_transactions = []

            for row_idx in range(self.start_row - 1, worksheet.nrows):
                try:
                    row_data = self._parse_excel_row(worksheet, row_idx)
                    if row_data:
                        # Tạo giao dịch
                        transaction = self._create_transaction(row_data)
                        if transaction:
                            created_transactions.append(transaction)
                            success_count += 1
                        else:
                            error_count += 1
                            error_messages.append(f"Dòng {row_idx + 1}: Không thể tạo giao dịch")

                except Exception as e:
                    error_count += 1
                    error_messages.append(f"Dòng {row_idx + 1}: {str(e)}")
                    _logger.error(f"Error importing row {row_idx + 1}: {str(e)}")

            # Cập nhật kết quả
            self.write({
                'import_status': 'done' if error_count == 0 else 'error',
                'success_count': success_count,
                'error_count': error_count,
                'error_log': '\n'.join(error_messages) if error_messages else '',
                'total_rows': worksheet.nrows - (self.start_row - 1)
            })

            # Hiển thị kết quả
            message = f"Import hoàn tất!\n"
            message += f"- Thành công: {success_count} giao dịch\n"
            message += f"- Lỗi: {error_count} giao dịch"

            if created_transactions:
                # Mở danh sách giao dịch vừa tạo
                return {
                    'type': 'ir.actions.act_window',
                    'name': 'Giao Dịch Đã Import',
                    'res_model': 'dpt.fund.transaction',
                    'view_mode': 'tree,form',
                    'domain': [('id', 'in', [t.id for t in created_transactions])],
                    'context': {'create': False}
                }
            else:
                return {
                    'type': 'ir.actions.act_window',
                    'res_model': self._name,
                    'res_id': self.id,
                    'view_mode': 'form',
                    'target': 'new',
                }

        except Exception as e:
            raise UserError(_('Lỗi import: %s') % str(e))

    def _parse_excel_row(self, worksheet, row_idx):
        """Parse một dòng từ Excel"""
        try:
            # Đọc các cột
            date_cell = worksheet.cell(row_idx, self.date_column - 1)
            description_cell = worksheet.cell(row_idx, self.description_column - 1)
            amount_cell = worksheet.cell(row_idx, self.amount_column - 1)
            reference_cell = worksheet.cell(row_idx, self.reference_column - 1) if self.reference_column else None
            balance_cell = worksheet.cell(row_idx, self.balance_column - 1) if self.balance_column else None

            # Parse ngày
            transaction_date = self._parse_date(date_cell)
            if not transaction_date:
                return None

            # Parse số tiền
            amount = self._parse_amount(amount_cell)
            if amount == 0:
                return None

            # Parse mô tả
            description = str(description_cell.value).strip() if description_cell.value else ''
            if not description:
                description = f'Giao dịch ngân hàng {transaction_date}'

            # Parse reference
            reference = str(reference_cell.value).strip() if reference_cell and reference_cell.value else ''

            # Parse balance
            balance = self._parse_amount(balance_cell) if balance_cell else 0

            return {
                'date': transaction_date,
                'description': description,
                'amount': amount,
                'reference': reference,
                'balance': balance
            }

        except Exception as e:
            raise ValidationError(f"Lỗi parse dòng {row_idx + 1}: {str(e)}")

    def _parse_date(self, cell):
        """Parse ngày từ Excel cell"""
        if not cell.value:
            return None

        try:
            if cell.ctype == xlrd.XL_CELL_DATE:
                # Excel date
                date_tuple = xlrd.xldate_as_tuple(cell.value, 0)
                return datetime(*date_tuple).date()
            else:
                # Text date
                date_str = str(cell.value).strip()
                if self.date_format == 'dd/mm/yyyy':
                    return datetime.strptime(date_str, '%d/%m/%Y').date()
                elif self.date_format == 'mm/dd/yyyy':
                    return datetime.strptime(date_str, '%m/%d/%Y').date()
                elif self.date_format == 'yyyy-mm-dd':
                    return datetime.strptime(date_str, '%Y-%m-%d').date()
                elif self.date_format == 'dd-mm-yyyy':
                    return datetime.strptime(date_str, '%d-%m-%Y').date()
        except:
            return None

    def _parse_amount(self, cell):
        """Parse số tiền từ Excel cell"""
        if not cell.value:
            return 0

        try:
            if isinstance(cell.value, (int, float)):
                return float(cell.value)
            else:
                # Remove common formatting
                amount_str = str(cell.value).replace(',', '').replace(' ', '')
                return float(amount_str)
        except:
            return 0

    def _create_transaction(self, row_data):
        """Tạo giao dịch từ dữ liệu dòng"""
        try:
            # Xác định loại giao dịch
            transaction_type = 'income' if row_data['amount'] > 0 else 'expense'

            # Tạo giao dịch
            transaction_vals = {
                'account_id': self.account_id.id,
                'date': row_data['date'],
                'description': row_data['description'],
                'amount': abs(row_data['amount']),
                'transaction_type': transaction_type,
                'reference': row_data['reference'],
                'state': 'confirmed',  # Tự động confirm giao dịch import
                'import_source': 'bank_excel',
            }

            # Kiểm tra trùng lặp
            if row_data['reference']:
                existing = self.env['dpt.fund.transaction'].search([
                    ('reference', '=', row_data['reference']),
                    ('account_id', '=', self.account_id.id)
                ])
                if existing:
                    _logger.warning(f"Duplicate transaction reference: {row_data['reference']}")
                    return None

            return self.env['dpt.fund.transaction'].create(transaction_vals)

        except Exception as e:
            _logger.error(f"Error creating transaction: {str(e)}")
            raise ValidationError(f"Không thể tạo giao dịch: {str(e)}")

    def action_download_template(self):
        """Tải template Excel mẫu"""
        # Tạo file Excel template
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Template')

        # Format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'color': 'white',
            'border': 1
        })

        data_format = workbook.add_format({
            'border': 1
        })

        # Headers
        headers = [
            'Ngày Giao Dịch',
            'Mô Tả',
            'Số Tiền',
            'Mã Tham Chiếu',
            'Số Dư'
        ]

        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # Sample data
        sample_data = [
            ['01/01/2024', 'Chuyển tiền lương', 15000000, 'REF001', 15000000],
            ['02/01/2024', 'Thanh toán điện nước', -500000, 'REF002', 14500000],
            ['03/01/2024', 'Thu tiền bán hàng', 2000000, 'REF003', 16500000],
        ]

        for row, data in enumerate(sample_data, 1):
            for col, value in enumerate(data):
                worksheet.write(row, col, value, data_format)

        # Thêm ghi chú
        worksheet.write(5, 0, "GHI CHÚ:", workbook.add_format({'bold': True}))
        worksheet.write(6, 0, "- Số tiền dương (+): Thu nhập")
        worksheet.write(7, 0, "- Số tiền âm (-): Chi phí")
        worksheet.write(8, 0, "- Định dạng ngày: DD/MM/YYYY")
        worksheet.write(9, 0, "- Mã tham chiếu giúp tránh trùng lặp")

        workbook.close()
        output.seek(0)

        # Tạo attachment
        template_data = base64.b64encode(output.read())
        attachment = self.env['ir.attachment'].create({
            'name': 'Template_Import_Giao_Dich_Ngan_Hang.xlsx',
            'type': 'binary',
            'datas': template_data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }
