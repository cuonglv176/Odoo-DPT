# -*- coding: utf-8 -*-
import base64
import io
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from odoo import models, fields, api, _


class DPTQuotationPrintWizard(models.TransientModel):
    _name = 'dpt.quotation.print.wizard'
    _description = 'Print Quotation Wizard'

    type = fields.Selection([
        ('quotation_total', 'Báo giá tổng'),
        ('quotation_product', 'Báo giá theo sản phẩm'),
        ('quotation_service', 'Báo giá theo dịch vụ tầng 1'),
    ], "Loại Form Báo giá", default="quotation_total")
    sale_order_id = fields.Many2one('sale.order', 'Sales Order')

    def action_confirm_print_quotation(self):
        file = self.get_excel_file()
        if file and file is not None:
            file_name = f"Form Báo giá {dict(self._fields['type'].selection).get(self.type)}.xlsx"
            attachment_id = self.env['ir.attachment'].search([('name', 'ilike', file_name)], limit=1)
            if attachment_id:
                attachment_id.unlink()
            attachment_id = self.env['ir.attachment'].sudo().create({
                'name': file_name,
                'type': 'binary',
                'datas': base64.b64encode(file),
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'res_model': 'clouds.folder',
            })
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment_id.id}?download=true',
                'target': 'self',
            }

    def get_excel_file(self):
        logo_path = "/Users/ungtu/Documents/Odoo/Freelancer/DPT_SOFT/Source/Odoo-DPT/dpt_print_quotation/static/description/logo.png"
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo giá dịch vụ"

        # Định nghĩa styles
        header_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=9)
        small_font = Font(name='Arial', size=8)
        title_font = Font(name='Arial', size=14, bold=True)
        company_font = Font(name='Arial', size=12, bold=True, color='FF6600')

        # Định nghĩa borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Alignment
        center_alignment = Alignment(horizontal='center', vertical='center')

        # === HEADER SECTION ===
        # Logo (A2:A5)
        ws.merge_cells('A2:A5')

        if logo_path and os.path.exists(logo_path):
            img = Image(logo_path)
            # Resize image if needed
            img.width = 150
            img.height = 60
            # Position the image in cell A1
            ws.add_image(img, 'A2')

        # Tiêu đề KY TỐC LOGISTICS (B2)
        ws['B2'] = 'KỲ TỐC LOGISTICS - 棋速'
        ws['B2'].font = Font(name='Arial', size=14, bold=True, color='FF6600')

        # Thông tin địa chỉ (B3)
        ws['B3'] = 'Địa chỉ: Số 6A, Ngõ 183, Hoàng Văn Thái, Khương Trung, Thanh Xuân, Hà Nội'
        ws['B3'].font = small_font

        # MST (B4)
        ws['B4'] = 'MST: 0109386059'
        ws['B4'].font = small_font

        # === TITLE SECTION ===
        # BÁO GIÁ DỊCH VỤ (Merge B6:D6)
        ws.merge_cells('B6:D6')
        ws['B6'] = 'BÁO GIÁ DỊCH VỤ'
        ws['B6'].font = title_font
        ws['B6'].alignment = center_alignment

        # Khách hàng (A7)
        ws['A7'] = 'Khách hàng:'
        ws['A7'].font = normal_font
        ws['B7'] = self.sale_order_id.partner_id.name or ""
        ws['B7'].font = normal_font

        # Địa chỉ (C7)
        ws['C7'] = 'Địa chỉ:'
        ws['C7'].font = normal_font

        ws['D7'] = self.sale_order_id.partner_id.street or ""
        ws['D7'].font = normal_font

        # Mặt hàng (A8)
        ws['A8'] = 'Mặt hàng:'
        ws['A8'].font = normal_font

        # Ghi chú (A9 merge to E9)
        ws.merge_cells('A9:E9')
        ws[
            'A9'] = 'Căn cứ quy định khách đã quan tâm tới dịch vụ của KY Tốc Logistics. Chúng tôi xin được gửi tới quý khách giá như sau:'
        ws['A9'].font = small_font

        # === TABLE HEADERS (Dòng 10) ===
        headers = [
            ('A10', 'Mã HS'),
            ('B10', ''),  # Trống cho cột B
            ('C10', 'Số lượng (cái)'),
            ('D10', 'Đơn giá (USD)'),
            ('E10', 'Note')
        ]

        for cell_ref, value in headers:
            cell = ws[cell_ref]
            cell.value = value
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # === TABLE DATA ===
        # Dòng 11 - Giá trị đầu tiên
        ws['A11'] = '0'
        ws['A11'].font = normal_font
        ws['A11'].border = thin_border
        ws['A11'].alignment = center_alignment

        # Border cho các ô trống trong dòng 11
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}11'].border = thin_border

        # === SECTION HÀNG HÓA ===
        # Dòng 12 - Header "Hàng hóa"
        ws.merge_cells('A12:A25')
        ws['A12'] = 'Hàng hóa'
        ws['A12'].font = header_font
        ws['A12'].border = thin_border

        # Border cho các ô khác trong dòng 12
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}12'].border = thin_border

        # Các mục trong section Hàng hóa
        hang_hoa_items = [
            (13, 'Thể tích (m3)', '', '', ''),
            (14, 'Khối lượng (kg)', '', '', ''),
            (15, 'Phi vận chuyển/ m3', 'VND', '', ''),
            (16, 'Phi vận chuyển/ kg', 'VND', '', ''),
            (17, 'NK CO Form E', '0%', '', ''),
            (18, 'VAT', '10%', '', ''),
            (19, '', 'Bảng Tương – HN', '', '')
        ]

        for row, item_b, item_c, item_d, item_e in hang_hoa_items:
            # Border cho tất cả các ô
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = thin_border

            # Điền dữ liệu
            if item_b:
                ws[f'B{row}'] = item_b
                ws[f'B{row}'].font = normal_font

            if item_c:
                ws[f'C{row}'] = item_c
                ws[f'C{row}'].font = normal_font
                ws[f'C{row}'].alignment = center_alignment

            if item_d:
                ws[f'D{row}'] = item_d
                ws[f'D{row}'].font = normal_font
                ws[f'D{row}'].alignment = center_alignment

            if item_e:
                ws[f'E{row}'] = item_e
                ws[f'E{row}'].font = normal_font

        # === CÁC KHOẢN PHÍ CHÍNH ===
        # Các mục với VND/lô và giá trị 0
        fee_items = [
            (20, 'Giá trị kê khai dự kiến', 'VND/lô', '0'),
            (21, 'Thuế NK', 'VND/lô', '0'),
            (22, 'Thuế VAT', 'VND/lô', '0'),
            (23, 'Phí ủy thác nhập khẩu', 'VND/lô', '0'),
            (24, 'Phí đấu mức', 'VND/lô', ''),
            (25, 'Phí nâng hạ', 'VND/lô', '')
        ]

        for row, item_name, unit, value in fee_items:
            # Border cho tất cả các ô
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = thin_border

            # Điền dữ liệu
            ws[f'B{row}'] = item_name
            ws[f'B{row}'].font = normal_font

            ws[f'C{row}'] = unit
            ws[f'C{row}'].font = normal_font

            if value:
                ws[f'D{row}'] = value
                ws[f'D{row}'].font = normal_font
                ws[f'D{row}'].alignment = center_alignment

        # === SECTION BÁO GIÁ CHI TIẾT ===
        # Dòng 26 - Header "Báo giá chi tiết"
        ws.merge_cells('A26:A33')
        ws['A26'] = 'Báo giá chi tiết'
        ws['A26'].font = header_font
        ws['A26'].border = thin_border

        # Border cho các ô khác trong dòng 26
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}26'].border = thin_border

        # Các mục trong báo giá chi tiết
        detail_items = [
            (27, 'Cước VC BT-HN (m3)', 'VND/lô', '0'),
            (28, 'Cước VC BT-HN (kg)', 'VND/lô', '0'),
            (29, 'Cước vận chuyển chăng cuốn (cây kiện)', 'VND/lô', '')
        ]

        for row, item_name, unit, value in detail_items:
            # Border cho tất cả các ô
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = thin_border

            # Điền dữ liệu
            ws[f'B{row}'] = item_name
            ws[f'B{row}'].font = normal_font

            ws[f'C{row}'] = unit
            ws[f'C{row}'].font = normal_font

            if value:
                ws[f'D{row}'] = value
                ws[f'D{row}'].font = normal_font
                ws[f'D{row}'].alignment = center_alignment

        # === SECTION TĂNG CHI PHÍ ===
        increase_items = [
            (30, 'Tăng chi phí vận chuyển theo m3', 'VND/lô', '0'),
            (31, 'Tăng chi phí vận chuyển theo kg', 'VND/lô', '0')
        ]

        for row, item_name, unit, value in increase_items:
            # Border cho tất cả các ô
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = thin_border

            # Điền dữ liệu
            ws[f'B{row}'] = item_name
            ws[f'B{row}'].font = normal_font

            ws[f'C{row}'] = unit
            ws[f'C{row}'].font = normal_font

            ws[f'D{row}'] = value
            ws[f'D{row}'].font = normal_font
            ws[f'D{row}'].alignment = center_alignment

        # === SECTION CHI PHÍ THEO ===
        cost_items = [
            (32, 'Chi phí theo m3', 'VND/m3', '#DIV/0!'),
            (33, 'Chi phí theo kg', 'VND/kg', '#DIV/0!')
        ]

        for row, item_name, unit, formula in cost_items:
            # Border cho tất cả các ô
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = thin_border

            # Điền dữ liệu
            ws[f'B{row}'] = item_name
            ws[f'B{row}'].font = normal_font

            ws[f'C{row}'] = unit
            ws[f'C{row}'].font = normal_font

            ws[f'D{row}'] = formula
            ws[f'D{row}'].font = normal_font

        # === FOOTER ===
        # Dòng 35 - Liên hệ
        ws['A35'] = 'Liên hệ: Chuyên viên:'
        ws['A35'].font = normal_font

        ws['D35'] = 'CÔNG TY TNHH DPT VINA HOLDINGS'
        ws['D35'].font = Font(name='Arial', size=10, bold=True)

        # Dòng 36 - SĐT
        ws['A36'] = 'SĐT:'
        ws['A36'].font = normal_font

        # Dòng 37 - Email
        ws['A37'] = 'Email:'
        ws['A37'].font = normal_font

        # === ĐIỀU CHỈNH KÍCH THƯỚC CỘT ===
        column_widths = [15, 35, 20, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Điều chỉnh chiều cao dòng
        for row in range(1, 40):
            ws.row_dimensions[row].height = 18

        # Save to BytesIO
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        return excel_bytes.getvalue()