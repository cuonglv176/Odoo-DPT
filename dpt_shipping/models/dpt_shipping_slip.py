# -*- coding: utf-8 -*-

import base64
from odoo import models, fields, api
from odoo.exceptions import ValidationError
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.drawing.image import Image
import io


class DPTShippingSlip(models.Model):
    _name = 'dpt.shipping.slip'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name = fields.Char('Name')
    active = fields.Boolean('Active', default=True)
    transfer_code = fields.Char('Transfer Code')
    transfer_code_chinese = fields.Char('Transfer Code in Chinese')
    main_in_picking_ids = fields.Many2many('stock.picking', 'stock_picking_main_incoming_shipping_rel',
                                           'shipping_slip_id', 'picking_id', string='Main In Picking',
                                           domain=[('state', '!=', 'cancel')])
    out_picking_ids = fields.Many2many('stock.picking', 'stock_picking_out_shipping_rel', 'shipping_slip_id',
                                       'picking_id', string='Out Picking', domain=[('state', '!=', 'cancel')])
    export_import_ids = fields.Many2many('dpt.export.import', 'export_import_shipping_rel', 'shipping_slip_id',
                                         'export_import_id', string='Export Import', store=True)
    export_import_name = fields.Char('Export Import Name', compute='_compute_information',
                                     search="_search_export_import_name")
    in_picking_ids = fields.Many2many('stock.picking', 'stock_picking_in_shipping_rel', 'shipping_slip_id',
                                      'picking_id', string='In Picking', domain=[('state', '!=', 'cancel')])
    sale_ids = fields.Many2many('sale.order', string='Sale Order', compute="_compute_information")
    ticket_ids = fields.Many2many('helpdesk.ticket', string="Tickets")
    vehicle_id = fields.Many2one('fleet.vehicle', 'Vehicle')
    vehicle_country = fields.Selection(related='vehicle_id.country')
    vn_vehicle_stage_id = fields.Many2one('dpt.vehicle.stage', 'Vietnamese Vehicle Stage',
                                          domain=[('country', '=', 'vietnamese')])
    cn_vehicle_stage_id = fields.Many2one('dpt.vehicle.stage', 'Chinese Vehicle Stage',
                                          domain=[('country', '=', 'chinese')])
    vehicle_stage_log_ids = fields.One2many('dpt.vehicle.stage.log', 'shipping_slip_id', 'Vehicle Stage Log')
    send_shipping_id = fields.Many2one('dpt.shipping.slip', 'Shipping Sent')
    vehicle_driver_id = fields.Many2one(related="vehicle_id.driver_id")
    vehicle_license_plate = fields.Char(related="vehicle_id.license_plate")
    vehicle_driver_phone = fields.Char(compute="_compute_vehicle_driver_phone")
    note = fields.Text('Note')

    total_volume = fields.Float('Total Volume (m3)', compute="_compute_information")
    total_weight = fields.Float('Total Weight (kg)', compute="_compute_information")
    total_num_packing = fields.Char('Total Num Packing', compute="_compute_information")
    num_not_confirm_picking = fields.Integer("Number of Not Confirm Picking", compute="_compute_information")
    estimate_arrival_warehouse_vn = fields.Date('Estimate Arrival Warehouse VN')
    non_finish_transfer = fields.Boolean('Non-Finish Transfer', compute="compute_non_finish_transfer")
    last_shipping_slip = fields.Boolean("Last Shipping Slip")
    is_cn_finish_stage = fields.Boolean(related="cn_vehicle_stage_id.is_finish_stage")
    is_vn_finish_stage = fields.Boolean(related="vn_vehicle_stage_id.is_finish_stage")
    all_so_locked = fields.Boolean("All SO Locked", compute="_compute_all_so_locked")
    delivery_slip_type = fields.Selection([
        ('container_tq', 'Container TQ'),
        ('container_vn', 'Container VN'),
        ('last_delivery_vn', 'Last Delivery VN'),
    ], "Delivery Slip Type")
    product_ids = fields.Many2many('product.product', 'Sản phẩm', compute="_compute_product")

    @api.depends('sale_ids', 'ticket_ids')
    def _compute_product(self):
        for item in self:
            product_ids = self.env['product.product']
            if item.sale_ids:
                product_ids |= item.sale_ids.mapped('order_line.product_id')
            if item.ticket_ids:
                product_ids |= item.ticket_ids.mapped('sale_id.order_line.product_id')
            item.product_ids = product_ids

    def _compute_all_so_locked(self):
        for item in self:
            item.all_so_locked = all([so_id.locked for so_id in item.sale_ids])

    def compute_non_finish_transfer(self):
        for item in self:
            item.non_finish_transfer = any([picking_id.total_left_quantity != 0 for picking_id in item.out_picking_ids])

    @api.constrains('estimate_arrival_warehouse_vn')
    def _constrains_get_arrival_warehouse_date(self):
        for item in self:
            if not item.send_shipping_id or not item.in_picking_ids:
                continue
            item.in_picking_ids.write({'estimate_arrival_warehouse_vn': self.estimate_arrival_warehouse_vn})

    def _compute_information(self):
        for item in self:
            item.sale_ids = (item.in_picking_ids | item.out_picking_ids).mapped(
                'sale_purchase_id') | item.export_import_ids.line_ids.mapped('sale_id')
            if item.vehicle_country == 'chinese' or not item.send_shipping_id:
                item.total_volume = sum(item.out_picking_ids.mapped('total_volume'))
                item.total_weight = sum(item.out_picking_ids.mapped('total_weight'))
                package_ids = item.out_picking_ids.mapped('package_ids')
            elif item.vehicle_country == 'vietnamese' or item.send_shipping_id:
                item.total_volume = sum(item.in_picking_ids.mapped('total_volume'))
                item.total_weight = sum(item.in_picking_ids.mapped('total_weight'))
                package_ids = item.in_picking_ids.mapped('package_ids')
            else:
                item.total_volume = 0
                item.total_weight = 0
                package_ids = None
            item.num_not_confirm_picking = len(
                item.out_picking_ids.filtered(lambda p: p.state != 'done')) if not item.send_shipping_id else len(
                item.in_picking_ids.filtered(lambda p: p.state != 'done'))

            item.export_import_name = ','.join(
                item.export_import_ids.filtered(lambda ei: ei.display_name).mapped('display_name'))

            # calculate total num packing
            total_num_packing_value = {}
            for package_id in package_ids:
                if not package_id.quantity:
                    continue
                if f'{package_id.uom_id.packing_code}' in total_num_packing_value:
                    total_num_packing_value[f'{package_id.uom_id.packing_code}'] = total_num_packing_value.get(
                        f'{package_id.uom_id.packing_code}') + package_id.quantity
                else:
                    total_num_packing_value[f'{package_id.uom_id.packing_code}'] = package_id.quantity
            item.total_num_packing = '.'.join([f"{quantity}{packing_code}" for packing_code, quantity in
                                               total_num_packing_value.items()]) if total_num_packing_value else None

    def _search_export_import_name(self, operator, value):
        return [('export_import_ids.display_name', operator, value)]

    @api.constrains('export_import_ids')
    @api.onchange('export_import_ids')
    def constrains_export_import(self):
        for item in self:
            sale_order_ids = item.export_import_ids.mapped('sale_id') | item.export_import_ids.mapped(
                'sale_ids') | item.export_import_ids.mapped('line_ids').mapped('sale_id')
            main_in_picking_ids = self.env['stock.picking'].search(
                [('sale_purchase_id', 'in', sale_order_ids.ids), ('is_main_incoming', '=', True)])
            # in_picking_ids = self.env['stock.picking'].search(
            #     [('sale_purchase_id', 'in', sale_order_ids.ids), ('x_transfer_type', '=', 'incoming_transfer')])
            out_picking_ids = self.env['stock.picking'].search(
                [('sale_purchase_id', 'in', sale_order_ids.ids),
                 ('x_transfer_type', '=', 'outgoing_transfer')]).filtered(
                lambda sp: not sp.out_shipping_ids or item.id in sp.out_shipping_ids.ids)
            item.sale_ids = [(6, 0, sale_order_ids.ids)]
            # item.in_picking_ids = [(6, 0, in_picking_ids.ids)]
            item.out_picking_ids = [(6, 0, out_picking_ids.ids)]
            item.main_in_picking_ids = [(6, 0, main_in_picking_ids.ids)]
            item.export_import_ids.shipping_slip_id = item.id

    def _compute_vehicle_driver_phone(self):
        for item in self:
            item.vehicle_driver_phone = item.vehicle_driver_id.phone or item.vehicle_driver_id.mobile if item.vehicle_driver_id else None

    @api.model
    def create(self, vals):
        if 'name' not in vals:
            vals['name'] = self._generate_service_code()
        return super(DPTShippingSlip, self).create(vals)

    def _generate_service_code(self):
        sequence = self.env['ir.sequence'].next_by_code('shipping.slip.seq') or '00'
        return f'{sequence}'

    def write(self, vals):
        if 'vn_vehicle_stage_id' not in vals and 'cn_vehicle_stage_id' not in vals:
            return super().write(vals)
        last_time_record_id = self.vehicle_stage_log_ids.sorted(key=lambda x: x['log_datetime'], reverse=True)[:1]
        last_time = last_time_record_id.log_datetime if last_time_record_id else self.create_date
        log_time = fields.Datetime.now() - last_time
        if self.vehicle_country == 'vietnamese':
            current_stage_id = self.vn_vehicle_stage_id
            res = super().write(vals)
            next_stage_id = self.vn_vehicle_stage_id
            if current_stage_id == next_stage_id:
                return res
            self.env['dpt.vehicle.stage.log'].create({
                'vehicle_id': self.vehicle_id.id,
                'shipping_slip_id': self.id,
                'current_stage_id': current_stage_id.id,
                'next_stage_id': next_stage_id.id,
                'log_time': log_time.total_seconds() / 3600,
            })
        elif self.vehicle_country == 'chinese':
            current_stage_id = self.cn_vehicle_stage_id
            res = super().write(vals)
            next_stage_id = self.cn_vehicle_stage_id
            if current_stage_id == next_stage_id:
                return res
            self.env['dpt.vehicle.stage.log'].create({
                'vehicle_id': self.vehicle_id.id,
                'shipping_slip_id': self.id,
                'current_stage_id': current_stage_id.id,
                'next_stage_id': next_stage_id.id,
                'log_time': log_time.total_seconds() / 3600,
            })
        else:
            return super().write(vals)
        return res

    @api.constrains('picking_ids', 'transfer_code', 'transfer_code_chinese')
    def constrains_transfer_code(self):
        for item in self:
            item.in_picking_ids.write({
                'transfer_code': item.transfer_code,
                'transfer_code_chinese': item.transfer_code_chinese,
            })

    def action_create_shipping_slip_receive(self):
        action = self.env.ref('dpt_shipping.dpt_shipping_split_wizard_action').sudo().read()[0]
        location_dest_id = self.env['stock.location'].sudo().search(
            [('usage', '=', 'internal'), ('warehouse_id.is_vn_transit_warehouse', '=', True)], limit=1)
        picking_ids = self.out_picking_ids.filtered(
            lambda p: p.state == 'done' and p.total_transfer_quantity)
        picking_ids.write({
            'delivery_dest_location_id': location_dest_id.id
        })
        action['context'] = {
            'default_shipping_id': self.id,
            'default_location_dest_id': location_dest_id.id,
            'default_picking_ids': picking_ids.ids,
            'default_sale_ids': picking_ids.mapped('sale_purchase_id').ids,
            'default_available_sale_ids': picking_ids.mapped('sale_purchase_id').ids,
            'default_available_picking_ids': picking_ids.ids,
        }
        return action

    def action_create_stock_transfer(self):
        ctq_location = self.env['stock.location'].sudo().search(
            [('warehouse_id.is_tq_transit_warehouse', '=', True), ('usage', '=', 'internal')], limit=1)
        if not ctq_location:
            raise ValidationError("Vui lòng kiểm tra lại kho chuyển phía Trung Quốc")
        for item in self:
            for main_incoming_picking_id in item.main_in_picking_ids:
                action = main_incoming_picking_id.action_create_transfer_picking()
                ctx = action['context']
                ctx.update({
                    'get_data_from_incoming': True,
                    'default_location_dest_id': ctq_location.id,
                })
                transfer_picking_id = self.env['stock.picking'].with_context(ctx).create({
                    'sale_purchase_id': main_incoming_picking_id.sale_purchase_id.id,
                })
                transfer_picking_id.action_update_old_package_information()
                transfer_picking_id._compute_total_volume_weight()

                # update move line
                move_line_vals = []
                for move_id in transfer_picking_id.move_ids_without_package:
                    move_id.move_line_ids.unlink()
                    move_id.location_dest_id = ctq_location
                    lot_id = self.env['stock.lot'].search(
                        [('product_id', '=', move_id.product_id.id),
                         ('name', '=', main_incoming_picking_id.picking_lot_name)],
                        limit=1)
                    move_line_vals.append({
                        'picking_id': transfer_picking_id.id,
                        'move_id': move_id.id,
                        'lot_id': lot_id.id,
                        'location_id': move_id.location_id.id,
                        'location_dest_id': ctq_location.id,
                        'product_id': move_id.product_id.id,
                        'quantity': move_id.product_uom_qty,
                        'product_uom_id': move_id.product_uom.id,
                    })
                if move_line_vals:
                    self.env['stock.move.line'].create(move_line_vals)
                # transfer_picking_id.create_in_transfer_picking()
            item.constrains_export_import()

    def action_confirm_picking(self):
        action = self.env.ref('dpt_shipping.dpt_picking_confirm_wizard_action').sudo().read()[0]
        picking_ids = self.out_picking_ids.filtered(
            lambda p: p.state != 'done') if not self.send_shipping_id else self.in_picking_ids.filtered(
            lambda p: p.state != 'done')
        action['context'] = {
            'default_shipping_id': self.id,
            'default_picking_ids': picking_ids.ids,
            'default_available_picking_ids': picking_ids.ids,
        }
        return action

    def action_lock_so(self):
        for sale_id in self.sale_ids.filtered(lambda so: not so.locked):
            sale_id.action_lock()
            sale_id.message_post(body="Đơn hàng bị khóa từ phiếu vận chuyển %s" % self.name,
                                 message_type='comment', subtype_xmlid='mail.mt_note')

    def action_export_picking_report(self):
        file = self.export_incoming_report_file()
        file_name = "Danh sách phiếu nhập kho.xlsx"
        if file_name and file:
            attachment_id = self.env['ir.attachment'].search([('name', 'ilike', file_name)])
            if attachment_id:
                attachment_id.unlink()
            attachment_id = self.env['ir.attachment'].sudo().create({
                'name': file_name,
                'type': 'binary',
                'datas': base64.b64encode(file),
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment_id.id}?download=true',
                'target': 'self',
            }

    def export_incoming_report_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Phiếu Nhập Kho"

        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 18

        # Set row heights
        ws.row_dimensions[1].height = 30  # Header row height
        ws.row_dimensions[2].height = 25  # Ngày nhập row
        ws.row_dimensions[3].height = 25  # Mã phiếu row
        ws.row_dimensions[4].height = 20  # Họ và tên người giao row
        ws.row_dimensions[5].height = 20  # Xuất tại kho row
        ws.row_dimensions[6].height = 20  # Nhập tại kho row
        ws.row_dimensions[7].height = 20  # Người nhận row
        ws.row_dimensions[8].height = 30  # Table header row

        # Set heights for table rows
        for row_num in range(9, 28):  # Table data rows
            ws.row_dimensions[row_num].height = 20

        # Define styles
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        dotted_border = Border(bottom=Side(style='dotted'))

        # Header
        # Add title
        ws.merge_cells('A1:I1')
        ws['A1'] = "PHIẾU NHẬP KHO"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        logo = Image(
            '/opt/odoo/odoo/Odoo-DPT-Test/dpt_stock_management/static/description/logo.png')
        ws.add_image(logo, 'B1')

        # Add mã id info on right top corner
        ws.merge_cells('J1:K1')
        ws['J1'] = "Mã id : 01 - VT"
        ws['J1'].font = Font(size=8)
        ws['J1'].alignment = Alignment(horizontal='center')
        ws['J1'].font = Font(bold=True)

        # Add QD info
        ws.merge_cells('J2:K3')
        ws['J2'] = "(Ban hành theo QĐ số: 48/2006/QĐ-BTC ngày 14/09/2006 của Bộ trưởng BTC)"
        ws['J2'].font = Font(size=8)
        ws['J2'].alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

        # Add date and form number
        ws['E2'] = "Ngày nhập:"
        ws['E2'].alignment = Alignment(horizontal='right')
        ws['F2'] = self.create_date.strftime("%d/%m/%Y")
        ws['F2'].alignment = Alignment(horizontal='right')
        ws['F2'].font = Font(bold=True)

        ws['E3'] = "Mã phiếu:"
        ws['E3'].alignment = Alignment(horizontal='right')
        ws['F3'] = self.name
        ws['F3'].alignment = Alignment(horizontal='right')
        ws['F3'].font = Font(bold=True)

        ws['E4'] = "Tên lái xe:"
        ws['E4'].alignment = Alignment(horizontal='right')
        ws['F4'] = self.vehicle_driver_id.name or ""
        ws['F4'].alignment = Alignment(horizontal='right')
        ws['F4'].font = Font(bold=True)

        ws['H4'] = "Biển số xe:"
        ws['H4'].alignment = Alignment(horizontal='left')
        ws['I4'] = self.vehicle_license_plate or ""
        ws['I4'].alignment = Alignment(horizontal='right')
        ws['I4'].font = Font(bold=True)

        # Add main information
        ws['B4'] = "Họ và tên người giao:"
        ws['B4'].alignment = Alignment(vertical='center')

        ws['B5'] = "Xuất tại kho:"
        ws['B5'].alignment = Alignment(vertical='center')
        ws['D5'] = self.in_picking_ids[0].location_id.warehouse_id.name or "" if self.in_picking_ids else ""
        ws['D5'].alignment = Alignment(horizontal='right')
        ws['D5'].border = dotted_border

        ws['B6'] = "Nhập tại kho :"
        ws['B6'].alignment = Alignment(vertical='center')
        ws['D6'] = ""
        ws['D6'].border = dotted_border

        ws['B7'] = "Người nhận:"
        ws['B7'].alignment = Alignment(vertical='center')

        # Add dotted underlines for text fields
        for row in range(4, 8):
            for col in range(ord('C'), ord('H')):
                if not ((row == 5 and col >= ord('D')) or (row == 6 and col >= ord('D'))):
                    cell = ws[f'{chr(col)}{row}']
                    cell.border = dotted_border

        # Table
        table_header = ["Số TT", "Khách hàng", "Mã đơn hàng", "Tên sản phẩm", "Mã lô", "Nhóm kiện", "Kích thước",
                        "Trọng lượng(kg)", "Thể tích(m3)", "Ghi chú"]
        for col, header in enumerate(table_header, 1):
            cell = ws.cell(row=9, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        datas = self.get_incoming_data()

        index = 1
        for data in datas:
            row10_data = [index, data.get("Khách hàng", ''), data.get("Mã đơn hàng", ''), data.get("Tên sản phẩm", ''),
                          data.get("Mã lô", ''), data.get("Nhóm kiện", ''), data.get("Kích thước", ''),
                          data.get("Trọng lượng(kg)", ''), data.get("Thể tích(m3)", ''), data.get("Thể tích(m3)", '')]
            for col, value in enumerate(row10_data, 1):
                cell = ws.cell(row=index + 9, column=col)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            index += 1

        # Instead of saving to file, save to BytesIO object
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)  # Move to the beginning of the BytesIO object

        # Return the bytes
        return excel_bytes.getvalue()

    def get_incoming_data(self):
        data = []
        if not self:
            return data
        for package_id in self.in_picking_ids.mapped('package_ids'):
            data.append({
                "Khách hàng": package_id.picking_id.sale_purchase_id.partner_id.name or "",
                "Mã đơn hàng": package_id.picking_id.sale_purchase_id.name or "",
                "Tên sản phẩm": ",".join(package_id.detail_ids.mapped('product_id.display_name')),
                "Mã lô": package_id.picking_id.picking_lot_name or "",
                "Nhóm kiện": f"{package_id.quantity}{package_id.uom_id.packing_code}",
                "Kích thước": f"{package_id.length}x{package_id.width}x{package_id.height}",
                "Trọng lượng(kg)": package_id.total_weight,
                "Thể tích(m3)": package_id.total_volume,
                "Ghi chú": "",
            })
        return data
