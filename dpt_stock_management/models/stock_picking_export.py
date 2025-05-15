# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import os
import io
import base64


class StockPicking(models.Model):
    _inherit = 'stock.picking'
    _rec_name = 'picking_lot_name'

    def action_export_picking_report(self):
        file = self.export_outgoing_report_file()
        file_name = "Danh sách phiếu xuất kho.xlsx"
        if file_name and file:
            attachment_id = self.env['ir.attachment'].search([('name', 'ilike', file_name)])
            if attachment_id:
                attachment_id.unlink()
            attachment_id = self.env['ir.attachment'].sudo().create({
                'name': file_name,
                'type': 'binary',
                'datas': base64.b64encode(file),
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment_id.id}?download=true',
                'target': 'self',
            }

    def export_outgoing_report_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Phiếu Xuất Kho"

        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 20

        # Define styles
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        dotted_border = Border(bottom=Side(style='dotted'))

        # Header
        # Add title
        ws.merge_cells('A1:I1')
        ws['A1'] = "PHIẾU XUẤT KHO"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        logo = Image(
            '/Users/ungtu/Documents/Odoo/Freelancer/DPT_SOFT/Source/Odoo-DPT/dpt_stock_management/static/description/logo.png')
        ws.add_image(logo, 'B1')

        # Add bill number
        ws.merge_cells('F3:G3')
        ws['F3'] = "Số phiếu:"
        ws['H3'] = self.name
        ws['H3'].font = Font(bold=True)

        # Add main information
        ws['B5'] = "Họ và tên người nhận:"
        ws['D5'] = self.partner_id.name or ""

        ws.merge_cells('F3:G3')
        ws['F5'] = "mã KH"
        ws['H5'] = self.partner_id.name or ""
        ws['H5'].font = Font(bold=True)

        ws['B6'] = "Địa chỉ :"
        ws['D6'] = self.partner_id.street or ""

        ws['B7'] = "Số điện thoại liên hệ:"
        ws['D7'] = self.partner_id.phone or ""

        ws['B8'] = "Xuất tại kho :"
        ws['D8'] = self.location_id.warehouse_id.name or ""
        for row in range(5, 9):
            for cell in ['B', 'G']:
                ws[f'{cell}{row}'].alignment = Alignment(vertical='center')

            # Add dotted underline for value cells
            for col in range(ord('C'), ord('I')):
                cell = ws[f'{chr(col)}{row}']
                cell.border = dotted_border

        # Table
        table_header = ["Số TT", "Mã đơn (ID)", "Mã Lô", "Sản phẩm", "Nhóm kiện", "Số kiện xuất", "Trọng lượng(kg)",
                        "Thể tích(m3)", "Ghi chú"]
        for col, header in enumerate(table_header, 1):
            cell = ws.cell(row=10, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        datas, total_quantity, total_weight, total_volume = self.get_outgoing_data()
        index = 1
        for data in datas:
            row10_data = [index, data.get("Mã đơn (ID)", ''), data.get("Mã Lô", ''), data.get("Sản phẩm", ''),
                         data.get("Nhóm kiện", ''), data.get("Số kiện xuất", ''), data.get("Trọng lượng(kg)", ''),
                         data.get("Thể tích(m3)", ''), data.get("Ghi chú", '')]
            for col, value in enumerate(row10_data, 1):
                cell = ws.cell(row=index + 10, column=col)
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            index += 1

        # Add total row
        row = index + 10
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=3).value = "Tổng"
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=6).value = total_quantity
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=7).value = total_weight
        ws.cell(row=row, column=7).border = thin_border
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=8).value = total_volume
        ws.cell(row=row, column=8).border = thin_border
        ws.cell(row=row, column=8).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=9).border = thin_border

        # Add note section
        ws.merge_cells(f'B{row + 3}:I{row + 3}')
        ws[f'B{row + 3}'] = """'+ Quý khách vui lòng kiểm tra số lượng và tình trạng kiện hàng trước khi ký nhận.Chúng tôi không tiếp nhận các khiếu nại không nhận được mã kiện có trong phiếu xuất kho này sau khi Quý khách đã ký nhận.\n+ Chúng tôi không giải quyết khiếu nại được (po sau 24h kể từ khi Khách hàng nhận hàng thành công.\n+ Chúng tôi không chịu bất kỳ trách nhiệm nào về hàng hóa (Bao gồm việc mất hàng hóa, móp méo, vỡ hỏng,...) sau khi giao hàng cho Nhà xe/Chành xe, Người nhận hộ."""
        ws[f'B{row + 3}'].alignment = Alignment(wrap_text=True)
        ws[f'B{row + 3}'].font = Font(bold=True)
        ws.row_dimensions[row + 3].height = 75

        # Add signature section
        ws[f'B{row + 6}'] = "NGƯỜI LẬP PHIẾU"
        ws[f'B{row + 6}'].alignment = Alignment(horizontal='center')
        ws[f'B{row + 6}'].font = Font(bold=True)

        ws[f'D{row + 6}'] = "THỦ KHO"
        ws[f'D{row + 6}'].alignment = Alignment(horizontal='center')
        ws[f'D{row + 6}'].font = Font(bold=True)

        ws[f'G{row + 6}'] = "NGƯỜI NHẬN HÀNG"
        ws[f'G{row + 6}'].alignment = Alignment(horizontal='center')
        ws[f'G{row + 6}'].font = Font(bold=True)
        ws.merge_cells(f'F{row + 7}:H{row + 7}')
        ws[f'F{row + 7}'] = "Nhận đủ số kiện và các kiện hàng nguyên"
        ws[f'F{row + 7}'].alignment = Alignment(horizontal='center')
        ws[f'F{row + 7}'].font = Font(bold=True)

        ws[f'I{row + 6}'] = "LÁI XE"
        ws[f'I{row + 6}'].alignment = Alignment(horizontal='center')
        ws[f'I{row + 6}'].font = Font(bold=True)

        # Add signature names
        ws[f'B{row + 10}'] = self.create_uid.name
        ws[f'B{row + 10}'].alignment = Alignment(horizontal='center')
        ws[f'B{row + 10}'].font = Font(bold=True)

        ws[f'D{row + 10}'] = "Nguyễn Đình Trường"
        ws[f'D{row + 10}'].alignment = Alignment(horizontal='center')
        ws[f'D{row + 10}'].font = Font(bold=True)

        # Instead of saving to file, save to BytesIO object
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)  # Move to the beginning of the BytesIO object

        # Return the bytes
        return excel_bytes.getvalue()

    def get_outgoing_data(self):
        data = []
        if not self:
            return data
        total_quantity = 0
        total_weight = 0
        total_volume = 0
        for package_id in self.package_ids:
            data.append({
                "Mã đơn(ID)": self.sale_purchase_id.name or "",
                "Mã Lô": self.picking_lot_name or "",
                "Sản phẩm": ",".join(package_id.detail_ids.mapped('product_id.display_name')),
                "Số kiện xuất": package_id.quantity,
                "Nhóm kiện": f"{package_id.quantity}{package_id.uom_id.packing_code}",
                "Trọng lượng(kg)": package_id.total_weight,
                "Thể tích(m3)": package_id.total_volume,
                "Ghi chú": "",
            })
            total_quantity += package_id.quantity
            total_weight += package_id.total_weight
            total_volume += package_id.total_volume
        return data, total_quantity, total_weight, total_volume

    def export_incoming_report_file(self):
        logo_path = "D:\Odoo17\Source\Odoo-DPT\dpt_stock_management\static\description\logo.png"
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Phiếu xuất kho"

        font_1 = Font(name='Times New Roman', size=24, bold=True)
        font_2 = Font(name='Times New Roman', size=12, bold=True, italic=True)
        font_3 = Font(name='Times New Roman', size=12, bold=True)

        # Define styles
        # Header title style - light blue
        title_fill = PatternFill(fill_type="solid")
        title_alignment = Alignment(horizontal='center', vertical='center')

        # Column header style - darker blue
        # header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        # header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # First merge horizontally for each row
        for row in range(2, 4):
            # Apply borders to all cells in the row
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        # Add company logo if provided
        if logo_path and os.path.exists(logo_path):
            img = Image(logo_path)
            # Position the image in cell A1
            ws.add_image(img, 'A2')

        ws.merge_cells('E2:I2')
        cell = ws.cell(row=2, column=4)
        cell.value = "PHIẾU NHẬP KHO"
        cell.font = font_1
        cell.alignment = title_alignment
        cell.fill = title_fill

        cell = ws.cell(row=2, column=10)
        cell.value = "Mẫu số : 01 - VT"
        cell.font = font_2
        cell.alignment = title_alignment
        cell.fill = title_fill

        # # Define column headers (row 9)
        # headers = [
        #     "No.", "문서번호", "보관부서", "기기번호\n(S/N)", "기기명\n(Description)",
        #     "형식\n(Model name)", "제작회사\n(Manufacturer)", "교정일자", "차기교정\n예정일",
        #     "관리주기 (월)", "교정대상", "교정기관", "용도", "기타", "교정유형", "검증\n여부", "사용\n여부"
        # ]
        #
        # # Add column headers
        # for col_idx, header in enumerate(headers, 1):
        #     cell = ws.cell(row=8, column=col_idx)
        #     cell.value = header
        #     cell.font = header_font
        #     cell.alignment = header_alignment
        #     cell.fill = header_fill
        #     cell.border = thin_border
        #
        # data_list = self.get_data()
        #
        # # Add sample data to worksheet
        # row_idx = 9  # Start from row 8 (after headers)
        # alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
        #
        # for idx, data in enumerate(data_list):
        #     # Apply alternating row colors
        #     row_fill = alt_row_fill if idx % 2 == 0 else None
        #
        #     for col_idx, header in enumerate(headers, 1):
        #         cell = ws.cell(row=row_idx, column=col_idx)
        #         cell.value = data.get(header, "")
        #         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        #         cell.border = thin_border
        #         if row_fill:
        #             cell.fill = row_fill
        #
        #     row_idx += 1
        #
        # # Set column widths
        # column_widths = {
        #     1: 5,  # No.
        #     2: 15,  # 문서번호
        #     3: 30,  # 보관부서
        #     4: 15,  # 기기번호
        #     5: 50,  # 기기명
        #     6: 30,  # 형식
        #     7: 30,  # 제작회사
        #     8: 12,  # 교정일자
        #     9: 12,  # 차기교정 예정일
        #     10: 12,  # 관리주기 (월)
        #     11: 10,  # 교정대상
        #     12: 50,  # 교정기관
        #     13: 25,  # 용도
        #     14: 50,  # 기타
        #     15: 20,  # 교정유형
        #     16: 8,  # 검증여부
        #     17: 8,  # 사용여부
        # }
        #
        # for col_num, width in column_widths.items():
        #     ws.column_dimensions[get_column_letter(col_num)].width = width

        # Instead of saving to file, save to BytesIO object
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)  # Move to the beginning of the BytesIO object

        # Return the bytes
        return excel_bytes.getvalue()